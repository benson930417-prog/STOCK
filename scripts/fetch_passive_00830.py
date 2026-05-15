import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


DATA_DIR = "data"
HISTORY_FILE = os.path.join(DATA_DIR, "passive_00830_history.json")
LOG_FILE = os.path.join(DATA_DIR, "passive_00830_log.json")
URL = "https://www.cathaysite.com.tw/ETF/detail/EBO?tab=etf3"
FALLBACK_URL = "https://www.etfinfo.tw/etf/00830/holdings"


def _num(value):
    text = str(value or "").replace(",", "").replace("$", "").replace("%", "").replace("新台幣", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _date_to_key(value):
    match = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", str(value or ""))
    if not match:
        raise ValueError(f"Could not parse 00830 date from {value!r}")
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"


def _load_json(path, default):
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def _write_json(path, payload):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
        fh.write("\n")


def _extract_page_meta(page):
    text = page.locator("body").inner_text(timeout=10000)
    if "Access Denied" in text or "edgesuite.net" in text:
        raise PermissionError("Cathay blocked this host with Access Denied")
    date_key = _date_to_key(text)

    def value_after(label):
        match = re.search(label + r".{0,40}?\$?\s*([0-9,]+(?:\.\d+)?)", text)
        return _num(match.group(1)) if match else None

    return date_key, {
        "fund_size": value_after("基金資產總淨值"),
        "nav": value_after("基金每單位淨值") or value_after("淨值"),
        "outstanding_units": int(value_after("基金在外流通單位數") or 0) or None,
        "closing_price": value_after("收盤價"),
        "source_url": URL,
    }


def _fallback_date_key(text):
    try:
        return _date_to_key(text)
    except ValueError:
        return datetime.now(timezone.utc).astimezone(ZoneInfo("Asia/Taipei")).date().isoformat()


def _download_excel(page):
    selectors = [
        "text=/匯出\\s*Excel/i",
        "text=/Excel/i",
        "a[href*='Excel']",
        "button:has-text('Excel')",
    ]
    for selector in selectors:
        locator = page.locator(selector).first
        try:
            if locator.count() == 0:
                continue
            with page.expect_download(timeout=15000) as download_info:
                locator.click()
            download = download_info.value
            path = Path(download.path())
            return path
        except Exception:
            continue
    raise ValueError("Could not trigger Cathay 00830 Excel download")


def _cell_text(value):
    return str(value or "").strip()


def _looks_like_holding_row(values):
    if len(values) < 4:
        return False
    code = _cell_text(values[0])
    return bool(re.match(r"^[A-Z0-9./-]{1,12}$", code, re.I)) and _num(values[-1]) is not None


def _parse_excel(path):
    wb = load_workbook(path, data_only=True)
    rows = []
    for ws in wb.worksheets:
        for raw in ws.iter_rows(values_only=True):
            values = [_cell_text(value) for value in raw if _cell_text(value)]
            if not _looks_like_holding_row(values):
                continue

            code = values[0].upper().replace(" ", "")
            name = values[1]
            nums = [_num(value) for value in values[2:]]
            nums = [value for value in nums if value is not None]
            if len(nums) < 2:
                continue

            shares = int(nums[0])
            weight = float(nums[-1])
            if weight <= 0 or weight > 100:
                continue
            if "." not in code and not code.endswith("US"):
                code = f"{code}.US"
            rows.append({
                "id": code,
                "name": name,
                "weight_pct": weight,
                "shares": shares,
            })

    # Deduplicate repeated print/header sections while preserving the first parsed instance.
    seen = set()
    holdings = []
    for row in rows:
        if row["id"] in seen:
            continue
        seen.add(row["id"])
        holdings.append(row)

    if len(holdings) < 10:
        raise ValueError(f"Only parsed {len(holdings)} 00830 holdings from Excel")
    return holdings


def _extract_etfinfo_holdings(page):
    text = page.locator("body").inner_text(timeout=15000)
    rows = page.evaluate(
        """() => {
            const rows = [];
            for (const tr of document.querySelectorAll('table tr')) {
                const cells = [...tr.querySelectorAll('th,td')]
                    .map(cell => cell.innerText.trim())
                    .filter(Boolean);
                if (cells.length >= 4) rows.push(cells);
            }
            return rows;
        }"""
    )

    holdings = []
    for values in rows:
        joined = " ".join(values)
        if "代號" in joined or "名稱" in joined:
            continue

        code = None
        code_index = None
        for index, value in enumerate(values):
            candidate = value.strip().upper().replace(" ", "")
            if re.match(r"^[A-Z]{1,6}(?:\.[A-Z]{1,4})?$", candidate):
                code = candidate
                code_index = index
                break
        if not code:
            continue

        weight = None
        for value in reversed(values):
            number = _num(value)
            if number is not None and 0 < number <= 100:
                weight = float(number)
                break
        if weight is None:
            continue

        shares = None
        for value in values:
            number = _num(value)
            if number is not None and number > 100:
                shares = int(number)
                break
        if shares is None:
            shares = 0

        name = values[code_index + 1] if code_index is not None and code_index + 1 < len(values) else code
        if "." not in code and not code.endswith("US"):
            code = f"{code}.US"
        holdings.append({
            "id": code,
            "name": name,
            "weight_pct": weight,
            "shares": shares,
        })

    deduped = []
    seen = set()
    for row in holdings:
        if row["id"] in seen:
            continue
        seen.add(row["id"])
        deduped.append(row)

    if len(deduped) < 10:
        raise ValueError(f"Only parsed {len(deduped)} fallback 00830 holdings from ETFInfo")
    return _fallback_date_key(text), {
        "fund_size": None,
        "nav": None,
        "outstanding_units": None,
        "closing_price": None,
        "source_url": FALLBACK_URL,
        "source_note": "Cathay page was blocked; holdings parsed from ETFInfo fallback.",
    }, deduped


def fetch_and_update_00830():
    now_utc = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    history = _load_json(HISTORY_FILE, {})
    previous_log = _load_json(LOG_FILE, {})

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(
            accept_downloads=True,
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
            ),
            locale="zh-TW",
            timezone_id="Asia/Taipei",
        )
        try:
            page.goto(URL, wait_until="networkidle", timeout=60000)
            page.wait_for_timeout(1000)
            date_key, meta = _extract_page_meta(page)
            excel_path = _download_excel(page)
            holdings = _parse_excel(excel_path)
        except PermissionError as exc:
            print(f"Cathay blocked 00830 fetch: {exc}. Falling back to ETFInfo.")
            page.goto(FALLBACK_URL, wait_until="networkidle", timeout=60000)
            page.wait_for_timeout(1000)
            date_key, meta, holdings = _extract_etfinfo_holdings(page)
        browser.close()

    payload = {
        "date": date_key,
        "meta": meta,
        "holdings": holdings,
    }
    previous = history.get(date_key)
    changed = previous != payload
    history[date_key] = payload
    _write_json(HISTORY_FILE, dict(sorted(history.items())))
    _write_json(
        LOG_FILE,
        {
            "last_checked_utc": now_utc,
            "last_updated_utc": now_utc if changed else previous_log.get("last_updated_utc"),
            "latest_date": date_key,
            "status": "NEW DATA FOUND" if changed else "NO CHANGE",
            "source": URL,
            "holdings_count": len(holdings),
        },
    )

    if changed:
        print(f"Successfully updated 00830 holdings for {date_key}. Total stocks: {len(holdings)}")
    else:
        print(f"No holding changes detected for 00830. Latest stored date remains {date_key}.")


if __name__ == "__main__":
    fetch_and_update_00830()
